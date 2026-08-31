"""XLSX engine for the unified document layer — branded layout.

Spreadsheets don't share the document :class:`ContentModel` (they have their
own sheets/rows/cols shape), so this engine consumes the native ``sheets``
payload **plus** a :class:`~kazma_core.documents.profile.DocProfile`. The
profile supplies the shared theme and direction, so an Arabic workbook matches
the Arabic DOCX/PDF/HTML/PPTX in design and direction.

Layout (mirrors the heading-bar motif of the other formats):
  - Row 1: branded title bar — merged across columns, accent fill, white bold.
  - Row 2: themed header band (heading fill, white bold).
  - Row 3+: data rows with alternating shading + grid.
  - Frozen title+header, RTL sheet view for Arabic, and print setup
    (landscape, fit-to-width, repeat title+header on every printed page).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from kazma_core.documents.profile import DocProfile

logger = logging.getLogger(__name__)

__all__ = ["XlsxEngine"]


class XlsxEngine:
    """Render a sheets payload to a branded ``.xlsx`` under a :class:`DocProfile`."""

    def __init__(self, profile: DocProfile) -> None:
        self.profile = profile
        self.theme = profile.theme

    def render(self, payload: dict[str, Any], output: Path | str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.properties import PageSetupProperties

        from kazma_core.documents.style_theme import theme_fonts

        t = self.theme
        rtl = self.profile.rtl
        font_name = theme_fonts(rtl=rtl)["cs"]
        accent_hex = str(t["accent"]).lstrip("#")
        header_hex = str(t["heading_fill"]).lstrip("#")
        body_color_hex = str(t["body"]).lstrip("#")
        grid_hex = str(t["table_grid"]).lstrip("#")
        alt_row_hex = str(t["table_row_bg"]).lstrip("#")

        accent_fill = PatternFill(fill_type="solid", fgColor=accent_hex)
        header_fill = PatternFill(fill_type="solid", fgColor=header_hex)
        alt_fill = PatternFill(fill_type="solid", fgColor=alt_row_hex)
        header_fg = str(t.get("table_header_fg") or "#16223a").lstrip("#")
        title_font = Font(bold=True, color="FFFFFF", name=font_name, size=14)
        header_font = Font(bold=True, color=header_fg, name=font_name, size=11)
        ink_fill = PatternFill(fill_type="solid", fgColor=str(t.get("heading") or "16223A").lstrip("#"))
        body_font = Font(name=font_name, color=body_color_hex, size=11)
        thin = Side(style="thin", color=grid_hex)
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(horizontal="right" if rtl else "left",
                          vertical="center", readingOrder=2 if rtl else 1)
        title_align = Alignment(horizontal="right" if rtl else "left",
                                vertical="center", readingOrder=2 if rtl else 1)

        workbook = Workbook()
        workbook.remove(workbook.active)

        # File core properties (title/author/subject).
        try:
            workbook.properties.title = str(payload.get("title") or "Workbook")
            workbook.properties.creator = str(payload.get("author") or "Kazma")
            if payload.get("subject"):
                workbook.properties.subject = str(payload.get("subject"))
            if payload.get("keywords"):
                workbook.properties.keywords = str(payload.get("keywords"))
        except Exception:
            logger.debug("[xlsx] core properties failed", exc_info=True)

        sheets = payload.get("sheets") if isinstance(payload.get("sheets"), list) else []
        for index, value in enumerate(sheets, 1):
            if not isinstance(value, dict):
                continue
            name = str(value.get("name", f"Sheet{index}"))[:31] or f"Sheet{index}"
            sheet = workbook.create_sheet(name)
            sheet.sheet_view.rightToLeft = rtl  # column A on the right for Arabic

            rows = [r for r in (value.get("rows") or []) if isinstance(r, list)]
            ncols = max((len(r) for r in rows), default=1) or 1

            # Row 1: branded title bar (merged across all columns).
            title_text = str(value.get("title") or name)
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            tcell = sheet.cell(row=1, column=1, value=title_text)
            tcell.fill = ink_fill
            tcell.font = title_font
            tcell.alignment = title_align
            sheet.row_dimensions[1].height = 28

            # Rows 2+: header (first data row) + body, all themed + bordered.
            for offset, row in enumerate(rows):
                sheet_row = 2 + offset
                is_header = (offset == 0)
                for c in range(1, ncols + 1):
                    cell = sheet.cell(row=sheet_row, column=c, value=self._cell(row[c - 1] if c <= len(row) else ""))
                    cell.font = header_font if is_header else body_font
                    if is_header:
                        cell.fill = header_fill
                    elif offset % 2 == 1:
                        cell.fill = alt_fill
                    cell.border = cell_border
                    cell.alignment = align

            # Reasonable column widths.
            from openpyxl.utils import get_column_letter

            for c in range(1, ncols + 1):
                length = max(
                    (len(str(sheet.cell(row=r, column=c).value or ""))
                     for r in range(2, len(rows) + 2)),
                    default=0,
                )
                col = get_column_letter(c)
                sheet.column_dimensions[col].width = min(48, max(10, length + 2))

            # Keep the title + header visible while scrolling.
            sheet.freeze_panes = "A3"

            # Print setup: landscape, fit to one page wide, repeat title+header.
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            sheet.print_title_rows = "1:2"

            # Optional chart (bar/line/pie) plotted over the sheet's data.
            chart_spec = value.get("chart")
            if isinstance(chart_spec, dict) and len(rows) >= 2:
                self._add_chart(sheet, chart_spec, last_data_row=len(rows) + 1)

        if not workbook.sheetnames:
            workbook.create_sheet("Sheet")
        workbook.save(str(output))

    @staticmethod
    def _add_chart(sheet: Any, spec: dict[str, Any], *, last_data_row: int) -> None:
        """Add a chart (bar/line/area/pie/doughnut/scatter) over the sheet's data.

        Layout assumption: row 1 = branded title, row 2 = header, rows 3+ = data.

        - ``type``: bar (vertical columns) | line | area | pie | doughnut | scatter.
        - ``category_col`` (default 1): the labels column.
        - ``value_col`` (default 2): a single column OR a list of columns for a
          multi-series chart (each header names its series).
        - ``title`` / ``x_axis`` / ``y_axis``: optional titles.
        """
        from openpyxl.chart import (AreaChart, BarChart, DoughnutChart, LineChart,
                                    PieChart, Reference, ScatterChart, Series)
        from openpyxl.chart.marker import Marker

        kind = str(spec.get("type", "bar")).lower()
        is_scatter = kind == "scatter"
        cls = {
            "bar": BarChart, "line": LineChart, "area": AreaChart,
            "pie": PieChart, "doughnut": DoughnutChart, "scatter": ScatterChart,
        }.get(kind, BarChart)
        chart = cls()
        chart.title = spec.get("title") or None
        if hasattr(chart, "x_axis"):
            chart.x_axis.title = spec.get("x_axis") or None
            chart.y_axis.title = spec.get("y_axis") or None
            chart.x_axis.majorGridlines = None  # cleaner look

        cat_col = int(spec.get("category_col", 1) or 1)
        raw_vcols = spec.get("value_col", 2)
        vcols = raw_vcols if isinstance(raw_vcols, list) else [int(raw_vcols or 2)]

        if is_scatter:
            # Scatter: x = category col, y = each value col (one Series each).
            x_ref = Reference(sheet, min_col=cat_col, min_row=3, max_row=last_data_row)
            for vc in vcols:
                y_ref = Reference(sheet, min_col=int(vc), min_row=3, max_row=last_data_row)
                series = Series(y_ref, x_ref)
                series.marker = Marker(symbol="circle", size=5)
                chart.series.append(series)
        else:
            for vc in vcols:
                data = Reference(sheet, min_col=int(vc), min_row=2, max_row=last_data_row)
                chart.add_data(data, titles_from_data=True)
            if kind not in {"pie", "doughnut"} and last_data_row >= 3 and cat_col <= sheet.max_column:
                cats = Reference(sheet, min_col=cat_col, min_row=3, max_row=last_data_row)
                chart.set_categories(cats)
        sheet.add_chart(chart, f"A{last_data_row + 2}")

    @staticmethod
    def _cell(item: Any) -> Any:
        """Normalise a payload cell: None -> empty string, keep numbers/dates."""
        if item is None:
            return ""
        return item
