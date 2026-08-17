"""Unification tests for the document layer (Stage 2).

These verify the architectural goal: the SAME payload produces a ContentModel +
DocProfile that BOTH the DOCX and PDF engines consume, and the direction /
alignment policy on the profile is the single place those decisions live. A
generated DOCX and a generated PDF are projections of one model under one
profile — one design, regardless of extension.
"""

from __future__ import annotations

from pathlib import Path

import pytest


def test_shared_builder_drives_both_formats(tmp_path: Path) -> None:
    """One payload → one (model, profile) → both DOCX and PDF engines.

    The DOCX and PDF generators must route through the shared
    ``_build_model_and_profile`` so a design change applies to both formats at
    once. This guards against either path re-growing format-specific payload
    parsing.
    """
    from kazma_core.documents.renderer_worker import _build_model_and_profile

    payload = {
        "title": "منظومة كاظمة",
        "subtitle": "تقرير",
        "lang": "ar",
        "toc": True,
        "sections": [{"heading": "المقدمة", "body": "نص عربي للفقرة الرئيسية هنا."}],
        "tables": [{"heading": "جدول", "headers": ["أ", "ب"], "rows": [["1", "2"]]}],
        "citations": ["مصدر أول"],
    }
    model, profile = _build_model_and_profile(payload)

    # The model carries every block type the payload expressed.
    from kazma_core.documents.content_model import (
        BodyBlock, CitationBlock, HeadingBlock, TableBlock, TitleBlock, TOCBlock,
    )
    block_types = [type(b) for b in model.blocks]
    assert TitleBlock in block_types          # title + subtitle
    assert TOCBlock in block_types
    assert HeadingBlock in block_types
    assert BodyBlock in block_types
    assert TableBlock in block_types
    assert CitationBlock in block_types

    # The profile is RTL for Arabic and carries the design + chrome.
    assert profile.rtl is True
    assert profile.chrome["toc"] == "المحتويات"

    # Both engines accept this exact (model, profile) and produce non-empty files.
    from kazma_core.documents.engines.docx import DocxEngine
    from kazma_core.documents.engines.pdf import PdfEngine

    docx_out = tmp_path / "out.docx"
    DocxEngine(profile).render(model, docx_out)
    assert docx_out.is_file() and docx_out.stat().st_size > 2000

    pdf_out = tmp_path / "out.pdf"
    PdfEngine(profile, []).render(model, pdf_out)
    assert pdf_out.is_file() and pdf_out.stat().st_size > 500


def test_profile_alignment_policy_is_direction_aware() -> None:
    """The alignment policy maps an *intent* to format-native values per direction.

    DOCX: "start" → "start" (bidi-logical; physical right under RTL). The
    bidi/jc inversion (jc=right → physical left) must never be produced.
    PDF:  "start" → "TA_RIGHT" (RTL) / "TA_LEFT" (LTR); "justify" → "TA_JUSTIFY".
    """
    from kazma_core.documents.profile import DocProfile

    ar = DocProfile.for_content("نص عربي", rtl=True)
    en = DocProfile.for_content("English text", rtl=False)

    # DOCX policy — uses the bidi-logical keyword "start" for BOTH directions
    # (start = physical left under LTR, physical right under RTL). "start" never
    # becomes "right" (which under w:bidi maps to the physical LEFT — the bug).
    assert ar.docx_jc("start") == "start"
    assert en.docx_jc("start") == "start"
    assert ar.docx_jc("justify") == en.docx_jc("justify") == "both"
    assert ar.docx_jc("end") == en.docx_jc("end") == "end"

    # PDF policy — visual mapping per direction (ReportLab is a visual engine).
    assert ar.pdf_align("start") == "TA_RIGHT"
    assert ar.pdf_align("justify") == "TA_JUSTIFY"
    assert en.pdf_align("start") == "TA_LEFT"
    assert en.pdf_align("justify") == "TA_JUSTIFY"


def test_shape_arabic_flag_reflects_content() -> None:
    """PDF engines pre-shape Arabic; the flag is RTL-or-Arabic-present.

    An LTR doc with embedded Arabic fragments must still shape (so glyphs join),
    matching the legacy ``shape_ar = rtl or is_arabic_dominant(sample)`` rule.
    """
    from kazma_core.documents.profile import DocProfile

    assert DocProfile.for_content("نص عربي واضح هنا", rtl=True).shape_arabic is True
    assert DocProfile.for_content("Plain English only", rtl=False).shape_arabic is False
    # LTR forced, but Arabic present → still shapes.
    assert DocProfile.for_content("مقدمة عربية", rtl=False).shape_arabic is True


def test_same_payload_yields_consistent_direction_across_overrides() -> None:
    """Explicit lang/rtl overrides resolve identically for both engines because
    they share one profile. Catches drift if either engine re-detects direction."""
    from kazma_core.documents.renderer_worker import _build_model_and_profile

    # Explicit rtl=False on Arabic content.
    _, prof1 = _build_model_and_profile({"title": "نص", "rtl": False})
    # Same via lang=ltr.
    _, prof2 = _build_model_and_profile({"title": "نص", "lang": "ltr"})
    assert prof1.direction == prof2.direction == "ltr"
    # And both still shape (Arabic present) — consistent PDF behavior.
    assert prof1.shape_arabic is True and prof2.shape_arabic is True


def test_html_engine_shares_profile_direction_theme_and_chrome(tmp_path: Path) -> None:
    """Stage 3: HTML export consumes the DocProfile, so an Arabic HTML doc
    carries dir=rtl + the shared THEME + localized chrome — matching the DOCX/PDF
    design. EN is dir=ltr. Guards against the old profile-free ``_markdown_html``
    regressing (it emitted no dir/lang/theme/chrome)."""
    from kazma_core.documents.renderer_worker import _generate_html
    from kazma_core.documents.style_theme import THEME

    ar = tmp_path / "ar.html"
    en = tmp_path / "en.html"
    payload_ar = {
        "title": "منظومة كاظمة", "lang": "ar", "toc": True,
        "sections": [{"heading": "المقدمة", "body": "نص عربي مع **تنسيق** و REST."}],
        "citations": ["مصدر أول"],
    }
    payload_en = {
        "title": "Kazma Platform", "lang": "en", "toc": True,
        "sections": [{"heading": "Intro", "body": "English body with **bold**."}],
        "citations": ["First source"],
    }
    _generate_html(ar, payload_ar)
    _generate_html(en, payload_en)
    ar_html = ar.read_text(encoding="utf-8")
    en_html = en.read_text(encoding="utf-8")

    # Direction + language from the profile.
    assert '<html lang="ar" dir="rtl">' in ar_html
    assert '<html lang="en" dir="ltr">' in en_html

    # Shared THEME token present in both (one design language).
    assert THEME["heading_fill"].lstrip("#") in ar_html
    assert THEME["heading_fill"].lstrip("#") in en_html

    # Localized chrome: Arabic TOC heading + references, English equivalents.
    assert "المحتويات" in ar_html
    assert "المراجع" in ar_html
    assert "Contents" in en_html
    assert "References" in en_html


def test_markdown_uses_localized_chrome_labels() -> None:
    """The Markdown emitter localizes Contents/References via the profile, so an
    Arabic .md matches the HTML/DOCX/PDF chrome (not hardcoded English)."""
    from kazma_core.documents.renderer_worker import _markdown

    ar_md = _markdown({"title": "تقرير", "lang": "ar", "toc": True,
                       "sections": [{"heading": "قسم", "body": "نص."}], "citations": ["مصدر"]})
    en_md = _markdown({"title": "Report", "toc": True,
                       "sections": [{"heading": "Sec", "body": "text."}], "citations": ["src"]})
    assert "## المحتويات" in ar_md and "## المراجع" in ar_md
    assert "## Contents" in en_md and "## References" in en_md


def test_convert_markdown_html_is_direction_aware() -> None:
    """convert:markdown:html (raw Markdown source, no payload) builds a profile
    from the text, so an Arabic Markdown source renders as dir=rtl HTML."""
    from kazma_core.documents.renderer_worker import _markdown_html

    ar_html = _markdown_html("# عنوان\n\nنص عربي واضح هنا في الفقرة الرئيسية للوثيقة.")
    en_html = _markdown_html("# Title\n\nPlain English markdown document body.")
    assert 'dir="rtl"' in ar_html and 'lang="ar"' in ar_html
    assert 'dir="ltr"' in en_html and 'lang="en"' in en_html


def test_xlsx_engine_is_themed_and_direction_aware(tmp_path: Path) -> None:
    """Stage 4: XLSX consumes the DocProfile — Arabic sheet view is right-to-left
    and header cells carry the shared theme fill; English is LTR. Both themed."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_xlsx

    rows = [["الميزة", "القيمة"], ["المعالجة", "سريعة"], ["اللغة", "عربية"]]
    ar = tmp_path / "ar.xlsx"
    en = tmp_path / "en.xlsx"
    _generate_xlsx(ar, {"title": "تقرير", "lang": "ar",
                        "sheets": [{"name": "البيانات", "rows": rows}]})
    _generate_xlsx(en, {"title": "Report", "lang": "en",
                        "sheets": [{"name": "Data", "rows": [["Feature", "Value"], ["Speed", "Fast"]]}]})

    def sheet_xml(p: Path) -> str:
        return zipfile.ZipFile(p).read("xl/worksheets/sheet1.xml").decode()

    def styles_xml(p: Path) -> str:
        return zipfile.ZipFile(p).read("xl/styles.xml").decode()

    # RTL sheet view for Arabic (rightToLeft="1"), not active for English.
    # openpyxl serializes rightToLeft="0" when False, so check the value.
    assert 'rightToLeft="1"' in sheet_xml(ar), "AR xlsx missing active rightToLeft sheet view"
    assert 'rightToLeft="1"' not in sheet_xml(en), "EN xlsx unexpectedly right-to-left"

    # Shared theme fill present in both (header / ink colour).
    from kazma_core.documents.style_theme import THEME

    ink = str(THEME["heading"]).lstrip("#").lower()
    assert ink in styles_xml(ar).lower()
    assert ink in styles_xml(en).lower()

    # Branded title row (row 1) merged across columns — the heading-bar motif
    # for spreadsheets, matching DOCX/PDF/HTML/PPTX.
    assert "mergeCell" in sheet_xml(ar), "XLSX missing merged branded title row"


def test_pptx_engine_is_themed_and_direction_aware(tmp_path: Path) -> None:
    """Stage 4: PPTX consumes the DocProfile — Arabic paragraphs carry rtl=1 and
    the title slide uses the shared accent/heading colour; English has no rtl."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_pptx

    ar = tmp_path / "ar.pptx"
    en = tmp_path / "en.pptx"
    _generate_pptx(ar, {"title": "منظومة كاظمة", "lang": "ar",
                        "slides": [{"heading": "المقدمة", "bullets": ["نقطة أولى", "نقطة ثانية"]}]})
    _generate_pptx(en, {"title": "Kazma", "slides": [{"heading": "Intro", "bullets": ["one", "two"]}]})

    def slide_xml(p: Path, n: int) -> str:
        return zipfile.ZipFile(p).read(f"ppt/slides/slide{n}.xml").decode()

    # Arabic content slide paragraphs are rtl=1 AND right-aligned (algn="r").
    # rtl alone does not right-align (inherits left from the layout); the
    # DrawingML alignment attr is "algn" (NOT "al" — that's silently ignored),
    # so algn="r" is required for Arabic to render right-aligned.
    ar_slide2 = slide_xml(ar, 2)
    assert 'rtl="1"' in ar_slide2, "AR pptx missing rtl paragraphs"
    assert 'algn="r"' in ar_slide2, "AR pptx missing right alignment (algn=r) on content"
    assert 'rtl="1"' not in slide_xml(en, 2), "EN pptx unexpectedly rtl"

    # 16:9 widescreen layout (modern default), not the legacy 4:3. Check by
    # ratio (cx/cy ≈ 1.78) — robust to python-pptx's exact EMU rounding.
    pres_xml = zipfile.ZipFile(ar).read("ppt/presentation.xml").decode()
    import re as _re
    m = _re.search(r'<p:sldSz\s+cx="(\d+)"\s+cy="(\d+)"', pres_xml)
    assert m, "PPTX presentation.xml missing sldSz"
    cx, cy = int(m.group(1)), int(m.group(2))
    assert cy == 6858000 and cx / cy > 1.6, (
        f"PPTX is not 16:9 (cx={cx}, cy={cy}, ratio={cx/cy:.2f})"
    )

    # Shared accent/heading colour on the title slide for both.
    from kazma_core.documents.style_theme import THEME

    ink = str(THEME["heading"]).lstrip("#").lower()
    accent = str(THEME["accent"]).lstrip("#").lower()
    ar1 = slide_xml(ar, 1).lower()
    en1 = slide_xml(en, 1).lower()
    assert ink in ar1 or accent in ar1
    assert ink in en1 or accent in en1


def test_pptx_speaker_notes_attached(tmp_path: Path) -> None:
    """PPTX speaker notes: a slide with a ``notes`` field gets a notes slide."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_pptx

    pptx = tmp_path / "deck.pptx"
    _generate_pptx(pptx, {
        "title": "Deck", "notes": "title notes",
        "slides": [{"heading": "Intro", "bullets": ["a"], "notes": "intro speaker notes"}],
    })
    names = zipfile.ZipFile(pptx).namelist()
    notes_xml = [n for n in names if "notesSlide" in n and n.endswith(".xml")]
    assert notes_xml, "PPTX missing notes slide"
    blob = "".join(zipfile.ZipFile(pptx).read(n).decode() for n in notes_xml)
    assert "intro speaker notes" in blob, "slide notes text not attached"


def test_docx_toc_field_and_page_numbers(tmp_path: Path) -> None:
    """DOCX emits a real TOC field (+ outline levels on section headings) and a
    PAGE field in the footer when ``page_numbers`` is set."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_docx

    docx = tmp_path / "t.docx"
    _generate_docx(docx, {
        "title": "Report", "lang": "en", "toc": True, "page_numbers": True,
        "sections": [{"heading": "Intro", "body": "body"}, {"heading": "End", "body": "body"}],
    })
    z = zipfile.ZipFile(docx)
    doc_xml = z.read("word/document.xml").decode()
    # TOC field instruction present.
    assert "TOC" in doc_xml and 'fldCharType="begin"' in doc_xml, "missing TOC field"
    # Section headings carry outline levels (indexable by the TOC field); the
    # TOC/References bars do NOT (so they don't self-list).
    assert doc_xml.count("<w:outlineLvl") == 2, "outline levels not on exactly the section headings"
    # Footer carries a PAGE field.
    footer = [n for n in z.namelist() if n.startswith("word/footer") and n.endswith(".xml")][0]
    fxml = z.read(footer).decode()
    assert "PAGE" in fxml and 'fldCharType="begin"' in fxml, "footer missing PAGE field"


def test_roundtrip_verifier_is_rtl_aware() -> None:
    """RTL round-trip: fuzzy token coverage (PyMuPDF glyph noise) then structural net.

    PyMuPDF returns logical-order Arabic with occasional glyph drops
    (االصطنا vs الاصطناعي). Exact substring can fail; token-set coverage must
    pass. Fully reversed pdfplumber output still passes the structural safety
    net so generation is not blocked if a fallback extractor is used. Empty /
    near-empty and LTR mismatches still fail.
    """
    from types import SimpleNamespace

    from kazma_core.documents.operations import DocumentOperations

    def doc(text: str) -> object:
        return SimpleNamespace(pages=[SimpleNamespace(blocks=[SimpleNamespace(text=text)])])

    ar_title = "منظومة كاظمة للذكاء الاصطناعي"
    # Near-correct PyMuPDF extract (logical order, minor glyph noise).
    pymupdf_ish = "منظومة كاظمة للذكاء االصطنا\nمقدمة\nنص عربي"
    # Fully reversed pdfplumber-style extract (safety-net path).
    mangled = "عيانطصلاا ءاكذلل ةمظاك ةموظنم\n(cid:1)\nالمقدمة\nنص عربي" * 3

    def passes(payload: dict, text: str) -> bool:
        try:
            DocumentOperations._validate_roundtrip_content("generate:pdf", payload, doc(text))
            return True
        except Exception:
            return False

    assert passes({"title": ar_title}, pymupdf_ish), "logical-order + glyph noise must pass via tokens"
    assert DocumentOperations._rtl_token_coverage(ar_title, pymupdf_ish) >= 0.5
    assert passes({"title": ar_title}, mangled), "reversed extract still passes structural net"
    assert passes({"title": "Kazma Report"}, "Kazma Report\nbody"), "EN exact match must pass"
    assert not passes({"title": "Kazma Report"}, "totally different"), "EN mismatch must fail"
    assert not passes({"title": ar_title}, "   \n   "), "Arabic empty/corrupt must fail"
    assert not passes({"title": ar_title}, "abc"), "Arabic near-empty (<20 non-ws) must fail"


def test_docx_update_fields_and_core_properties(tmp_path: Path) -> None:
    """Tier A: DOCX sets updateFields (TOC/page# auto-populate on open) and
    writes title/author/subject into the file's core properties."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_docx

    docx = tmp_path / "t.docx"
    _generate_docx(docx, {
        "title": "Quarterly Report", "author": "Alice", "subject": "Finance",
        "keywords": "q3, 2026", "lang": "en", "toc": True,
        "sections": [{"heading": "Overview", "body": "body"}],
    })
    z = zipfile.ZipFile(docx)
    settings = z.read("word/settings.xml").decode()
    assert "updateFields" in settings, "DOCX missing updateFields (fields won't auto-update on open)"
    core = z.read("docProps/core.xml").decode()
    assert "Quarterly Report" in core and "Alice" in core and "Finance" in core, (
        "DOCX core properties not populated"
    )


def test_image_embedding_approved_only(tmp_path: Path) -> None:
    """Batch 2: an approved-asset image is embedded; a missing/unapproved one is
    silently skipped — the security property (only sha256-validated assets reach
    the document) holds across DOCX/HTML/MD."""
    import zipfile

    pytest.importorskip("PIL")
    import io

    from PIL import Image as PILImage
    from kazma_core.documents.renderer_worker import _generate_docx, _generate_html, _markdown

    assets = tmp_path / "assets"
    assets.mkdir()
    buf = io.BytesIO()
    PILImage.new("RGB", (120, 80), (30, 58, 95)).save(buf, "PNG")
    (assets / "logo.png").write_bytes(buf.getvalue())

    payload = {
        "title": "Report", "lang": "en",
        "sections": [{"heading": "Intro", "body": "text"}],
        "images": [
            {"name": "logo.png", "caption": "approved", "width_in": 2.0},
            {"name": "notapproved.png", "caption": "should be skipped"},
        ],
    }

    # DOCX: approved image embedded; unapproved absent.
    docx = tmp_path / "out.docx"
    _generate_docx(docx, payload, assets_dir=assets)
    media = [n for n in zipfile.ZipFile(docx).namelist() if n.startswith("word/media/")]
    assert media, "approved image not embedded in DOCX"
    assert len(media) == 1, f"unapproved image leaked into DOCX: {media}"

    # HTML: approved image as a data-URI; unapproved not referenced.
    html = tmp_path / "out.html"
    _generate_html(html, payload, assets_dir=assets)
    h = html.read_text(encoding="utf-8")
    assert "data:image/png;base64," in h, "approved image not embedded in HTML"
    assert "notapproved.png" not in h, "unapproved image leaked into HTML"

    # Markdown: references both by name (MD can't embed binaries), but that's a
    # text reference, not an embed — acceptable.
    md = _markdown(payload)
    assert "![approved](logo.png)" in md


def test_office_core_properties_xlsx_pptx(tmp_path: Path) -> None:
    """Tier A: XLSX and PPTX populate title/author/subject core properties."""
    import zipfile

    from kazma_core.documents.renderer_worker import _generate_pptx, _generate_xlsx

    xlsx = tmp_path / "t.xlsx"
    pptx = tmp_path / "t.pptx"
    _generate_xlsx(xlsx, {"title": "Sales Data", "author": "Bob", "subject": "Q3",
                          "lang": "en", "sheets": [{"name": "S", "rows": [["a", "b"]]}]})
    _generate_pptx(pptx, {"title": "Deck", "author": "Carol", "subject": "Subj",
                          "slides": [{"heading": "S", "bullets": ["x"]}]})
    x_core = zipfile.ZipFile(xlsx).read("docProps/core.xml").decode()
    p_core = zipfile.ZipFile(pptx).read("docProps/core.xml").decode()
    assert "Sales Data" in x_core and "Bob" in x_core and "Q3" in x_core
    assert "Deck" in p_core and "Carol" in p_core and "Subj" in p_core


def test_xlsx_formulas_and_charts(tmp_path: Path) -> None:
    """XLSX: cells starting with '=' are written as formulas; a ``chart`` spec
    adds a bar/line/pie chart over the sheet's data."""
    from openpyxl import load_workbook

    from kazma_core.documents.renderer_worker import _generate_xlsx

    xlsx = tmp_path / "t.xlsx"
    _generate_xlsx(xlsx, {
        "title": "Sales", "lang": "en",
        "sheets": [{
            "name": "Q",
            "rows": [["Month", "Sales", "Total"], ["Jan", 100, "=B3*2"], ["Feb", 200, "=B4*2"]],
            "chart": {"type": "bar", "title": "Sales by Month"},
        }],
    })
    wb = load_workbook(xlsx)
    ws = wb["Q"]
    # Formula cell preserved as a formula (data_type 'f'), not stringized.
    assert ws["C3"].value == "=B3*2" and ws["C3"].data_type == "f", (
        f"formula cell stringized: {ws['C3'].value!r} ({ws['C3'].data_type})"
    )
    # Chart added.
    assert len(ws._charts) >= 1, "chart not added to sheet"


def test_xlsx_chart_types_and_multiseries(tmp_path: Path) -> None:
    """Batch 3: bar/line/area/pie/doughnut/scatter chart types render, and a
    multi-series chart (value_col as a list) plots one series per column."""
    from openpyxl import load_workbook

    from kazma_core.documents.renderer_worker import _generate_xlsx

    rows = [["Month", "Sales", "Cost"], ["Jan", 100, 60], ["Feb", 200, 120], ["Mar", 150, 90]]
    for kind, vcol in [("bar", [2, 3]), ("line", [2, 3]), ("area", [2, 3]),
                       ("pie", [2, 3]), ("doughnut", [2, 3]), ("scatter", 2)]:
        xlsx = tmp_path / f"{kind}.xlsx"
        _generate_xlsx(xlsx, {"title": "T", "lang": "en",
                              "sheets": [{"name": "S", "rows": rows,
                                          "chart": {"type": kind, "value_col": vcol}}]})
        ws = load_workbook(xlsx)["S"]
        assert len(ws._charts) == 1, f"{kind}: chart not added"
    # Multi-series bar: value_col [2,3] → 2 series.
    multi = tmp_path / "multi.xlsx"
    _generate_xlsx(multi, {"title": "T", "lang": "en",
                           "sheets": [{"name": "S", "rows": rows,
                                       "chart": {"type": "bar", "value_col": [2, 3]}}]})
    ws = load_workbook(multi)["S"]
    assert len(ws._charts[0].series) == 2, "multi-series chart did not plot 2 series"


def test_pdf_toc_has_page_numbers(tmp_path: Path) -> None:
    """PDF real TOC: the TableOfContents flowable renders entries WITH page
    numbers (via multiBuild two-pass). The old flat list had no page numbers."""
    import pymupdf
    from kazma_core.documents.renderer_worker import _generate_pdf

    pdf = tmp_path / "toc.pdf"
    _generate_pdf(pdf, {
        "title": "Report", "lang": "en", "toc": True,
        "sections": [
            {"heading": "Introduction", "body": "Body text for the introduction."},
            {"heading": "Conclusion", "body": "Body text for the conclusion."},
        ],
    }, [])
    doc = pymupdf.open(str(pdf))
    page1 = doc[0].get_text()
    doc.close()
    # The TOC area should contain entry names AND bare digits (page numbers).
    assert "Introduction" in page1, "TOC missing entry text"
    assert "Conclusion" in page1, "TOC missing entry text"
    # Old flat-list format would be "1. Introduction"; the real TOC separates
    # entry text from page numbers. Check that at least one bare digit appears
    # (the page number from the TableOfContents flowable).
    assert any(c.isdigit() for c in page1), "TOC missing page numbers"


def test_html_code_highlighting(tmp_path: Path) -> None:
    """HTML code highlighting: fenced code blocks get Pygments syntax coloring."""
    from kazma_core.documents.renderer_worker import _generate_html

    html = tmp_path / "code.html"
    _generate_html(html, {
        "title": "Report", "lang": "en",
        "sections": [{"heading": "Demo", "body": "```python\ndef foo():\n    pass\n```"}],
    })
    h = html.read_text(encoding="utf-8")
    assert "codehilite" in h, "HTML missing codehilite class"
    assert "<span" in h, "HTML missing Pygments syntax-highlight spans"


def test_html_clickable_toc(tmp_path: Path) -> None:
    """HTML clickable TOC: entries are <a href="#slug"> links; headings have
    matching id="slug" attributes so clicking scrolls to the section."""
    from kazma_core.documents.renderer_worker import _generate_html

    html = tmp_path / "toc.html"
    _generate_html(html, {
        "title": "Report", "lang": "en", "toc": True,
        "sections": [
            {"heading": "Introduction", "body": "body"},
            {"heading": "Conclusion", "body": "body"},
        ],
    })
    h = html.read_text(encoding="utf-8").lower()
    assert 'href="#introduction"' in h, "TOC entries not clickable links"
    assert 'id="introduction"' in h, "Section headings missing anchor IDs"
