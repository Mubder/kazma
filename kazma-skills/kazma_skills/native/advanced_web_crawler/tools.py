"""Advanced Web Crawler Native Skill — tools for search, crawling, and document parsing."""

from __future__ import annotations

import logging
import json
import csv
from pathlib import Path
from typing import Any

from kazma_core.tools.web_search import web_search
from kazma_core.tools.read_url import read_url
from kazma_core.agent.tool_registry import _workspace_scope_error
from kazma_core.tools.file_write import _get_workspace

logger = logging.getLogger(__name__)


async def web_search_duckduckgo(query: str, limit: int = 5) -> str:
    """Search DuckDuckGo or Bing fallback and return markdown results.

    Args:
        query: The search query string.
        limit: Max number of results.

    Returns:
        A markdown-formatted string with search results.
    """
    try:
        return await web_search(query, max_results=limit)
    except Exception as e:
        logger.error("Error executing web search: %s", e)
        return f"Error executing web search: {e}"


async def crawl_page(url: str) -> str:
    """Fetch and extract clean markdown content from a public URL.

    Args:
        url: The URL of the web page to crawl.

    Returns:
        Extracted text or markdown content, or an error message.
    """
    try:
        return await read_url(url)
    except Exception as e:
        logger.error("Error crawling webpage: %s", e)
        return f"Error crawling webpage: {e}"


async def parse_document(path: str) -> str:
    """Parse structured text from local files including CSV, JSON, XLS/XLSX, and PDF.

    Args:
        path: Path to the local file to parse.

    Returns:
        Organized textual representation of the document contents.
    """
    p = Path(path).expanduser().resolve()
    scope_err = _workspace_scope_error(p, path, "reads")
    if scope_err:
        return scope_err

    if not p.exists():
        return f"Error: Document not found: {path}"
    if not p.is_file():
        return f"Error: Path is not a file: {path}"

    suffix = p.suffix.lower()

    try:
        if suffix == ".json":
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                data = json.load(f)
            return json.dumps(data, indent=2, ensure_ascii=False)[:8000]

        elif suffix == ".csv":
            rows = []
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 100:  # Cap at 100 rows
                        rows.append("[... truncated after 100 rows ...]")
                        break
                    rows.append(", ".join(row))
            return "\n".join(rows)

        elif suffix in (".xls", ".xlsx"):
            return _parse_excel(p)

        elif suffix == ".pdf":
            return _parse_pdf(p)

        elif suffix in (".txt", ".md", ".log"):
            content = p.read_text(encoding="utf-8", errors="replace")
            if len(content) > 8000:
                return content[:8000] + "\n[... truncated ...]"
            return content

        else:
            # Basic fallback for other formats
            content = p.read_text(encoding="utf-8", errors="replace")
            if len(content) > 4000:
                content = content[:4000] + "\n[... truncated ...]"
            return f"[Parsed as Text ({suffix or 'no extension'})]:\n{content}"

    except Exception as e:
        logger.error("Error parsing document %s: %s", path, e)
        return f"Error parsing document: {e}"


def _parse_excel(p: Path) -> str:
    """Parse XLS/XLSX files using openpyxl. Returns rows as text."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = load_workbook(p, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 100:
                lines.append("[... truncated after 100 rows ...]")
                break
            # Convert all cells to string, skip fully-empty rows
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                lines.append(" | ".join(cells))
        lines.append("")
    wb.close()
    result = "\n".join(lines)
    return result[:8000] + "\n[... truncated ...]" if len(result) > 8000 else result


def _parse_pdf(p: Path) -> str:
    """Parse PDF files using PyPDF2 or pdfplumber. Returns extracted text."""
    # Try pdfplumber first (better extraction quality)
    try:
        import pdfplumber

        lines = []
        with pdfplumber.open(p) as pdf:
            for i, page in enumerate(pdf.pages[:50]):  # Cap at 50 pages
                text = page.extract_text() or ""
                if text.strip():
                    lines.append(f"--- Page {i + 1} ---")
                    lines.append(text.strip())
                    lines.append("")
        result = "\n".join(lines)
        if result.strip():
            return result[:8000] + "\n[... truncated ...]" if len(result) > 8000 else result
    except ImportError:
        pass
    except Exception as e:
        logger.debug("pdfplumber failed: %s, trying PyPDF2", e)

    # Fallback: PyPDF2
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(p))
        lines = []
        for i, page in enumerate(reader.pages[:50]):
            text = page.extract_text() or ""
            if text.strip():
                lines.append(f"--- Page {i + 1} ---")
                lines.append(text.strip())
                lines.append("")
        result = "\n".join(lines)
        return result[:8000] + "\n[... truncated ...]" if len(result) > 8000 else result
    except ImportError:
        return "Error: No PDF library installed. Run: pip install pdfplumber"
    except Exception as e:
        return f"Error parsing PDF: {e}"
